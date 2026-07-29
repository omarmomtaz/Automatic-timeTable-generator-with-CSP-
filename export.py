from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from data import DAYS, PERIODS, SECTIONS

SESSION_FILL = {
    "LEC": PatternFill("solid", fgColor="D9E7F5"),   # light blue
    "TUT": PatternFill("solid", fgColor="E2F0D9"),   # light green
    "LAB": PatternFill("solid", fgColor="FDEBD0"),   # light orange
}
HEADER_FILL = PatternFill("solid", fgColor="2E4057")
DAY_FILL = PatternFill("solid", fgColor="BFBFBF")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cell_text(v):
    return f"{v.code} - {v.name}\n[{v.session}]  {v.instructor}"


def export_semester(semester, assignment, by_id, out_path):
    # organize: by year -> section -> (day,period) -> Variable+room
    grid = {y: {s: {} for s in SECTIONS[y]} for y in (1, 2, 3, 4)}
    for vid, (day, period, room) in assignment.items():
        v = by_id[vid]
        for sec in v.sections:
            year = int(sec.split("-")[0][1])
            if sec not in grid[year]:
                continue
            grid[year][sec][(day, period)] = (v, room)

    wb = Workbook()
    wb.remove(wb.active)

    for year in (1, 2, 3, 4):
        sections = SECTIONS[year]
        ws = wb.create_sheet(f"Year {year} - {semester}")
        ws.sheet_view.showGridLines = False

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(sections))
        title = ws.cell(row=1, column=1, value=f"CSIT  |  Year {year}  |  {semester} Semester Timetable")
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = HEADER_FILL
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Header row
        header_row = 2
        ws.cell(row=header_row, column=1, value="Day").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=header_row, column=1).fill = HEADER_FILL
        ws.cell(row=header_row, column=2, value="Time").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=header_row, column=2).fill = HEADER_FILL
        for j, sec in enumerate(sections):
            c = ws.cell(row=header_row, column=3 + j, value=sec)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")

        row = header_row + 1
        for d_idx, day in enumerate(DAYS):
            start_row = row
            for p_idx, period in enumerate(PERIODS):
                ws.cell(row=row, column=2, value=period).alignment = Alignment(horizontal="center")
                for j, sec in enumerate(sections):
                    cell = ws.cell(row=row, column=3 + j)
                    entry = grid[year][sec].get((d_idx, p_idx))
                    if entry:
                        v, room = entry
                        cell.value = f"{_cell_text(v)}\nRoom: {room}"
                        cell.fill = SESSION_FILL.get(v.session, PatternFill())
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                    cell.border = BORDER
                ws.cell(row=row, column=2).border = BORDER
                ws.row_dimensions[row].height = 46
                row += 1
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
            dcell = ws.cell(row=start_row, column=1, value=day)
            dcell.font = Font(bold=True)
            dcell.fill = DAY_FILL
            dcell.alignment = Alignment(horizontal="center", vertical="center")
            dcell.border = BORDER

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        for j in range(len(sections)):
            ws.column_dimensions[get_column_letter(3 + j)].width = 30

        ws.freeze_panes = "C3"

    wb.save(out_path)
